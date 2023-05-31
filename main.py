import os
import csv
import shutil
import openpyxl
from tkinter import filedialog, Tk, Button, Label, Entry


input_file = ""
directory_to_search = "S:\\Consumer Files"
#directory_to_search = "C:\\Users\\mmullins\\OneDrive - Center for Independent Living in Central Florida\\Desktop\\from"
#destination_folder = "C:\\Users\\mmullins\\OneDrive - Center for Independent Living in Central Florida\\Desktop\\to"
destination_folder = "C:\\Users\\mmullins\\OneDrive - Center for Independent Living in Central Florida\\records\\General"
output_csv = os.path.join(destination_folder, "notFound.csv")
print(f"Output CSV: {output_csv}")
keywords = ["ROI", "ELIG DETERM", "ILP"]


def read_excel(file_path, id_column, last_name_column, start_row):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    folder_names = []
    for row in range(start_row, ws.max_row + 1):
        id_value = ws.cell(row=row, column=id_column).value
        last_name_value = ws.cell(row=row, column=last_name_column).value
        if id_value and last_name_value:
            # switch the position of ID and Last Name
            folder_name = f"{last_name_value.strip()} {id_value.strip()}"
            folder_names.append(folder_name)
    return folder_names



def column_to_index(column):
    """
    Convert an Excel column letter into a 1-indexed numerical column index.
    """
    number = 0
    for i, letter in enumerate(reversed(column.upper())):
        number += (ord(letter) - 64) * (26 ** i)
    return number


def find_nonexistent_folders(folder_names, search_directory):
    not_found_folders = []
    for folder_name in folder_names:
        folder_name = folder_name.strip("\\")
        folder_path = os.path.join(search_directory, folder_name)
        if not os.path.exists(folder_path):
            not_found_folders.append(folder_name)
    return not_found_folders


def read_csv(file_path):
    folder_names = []
    with open(file_path, newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            folder_names.append(row[0])
    return folder_names


def check_folders_exist(folder_names, directory):
    print("Checking Folders")
    not_found_folders = []
    for folder_name in folder_names:
        folder_path = os.path.join(directory, folder_name)
        if not os.path.exists(folder_path):
            not_found_folders.append(folder_name)
    return not_found_folders


def write_not_found_csv(not_found_folders, output_csv):
    file_exists = os.path.exists(output_csv)
    with open(output_csv, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        if not file_exists:
            writer.writerow(['Folder Name'])
        for folder_name in not_found_folders:
            writer.writerow([folder_name])


def write_to_csv(not_found_list, output_file):
    try:
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Folder Name'])
            for folder in not_found_list:
                writer.writerow([folder])
        print(f"Data written to {output_file} successfully.")
    except Exception as e:
        print(f"Error writing to file: {output_file}. Error: {e}")


def copy_files_with_keywords(folder_names, search_directory, target_directory, keywords):
    with open('log.csv', 'w', newline='') as log_file:
        csv_writer = csv.writer(log_file)
        csv_writer.writerow(['Folder Name', 'Path Copied From', 'Path Pasted To'])

        for folder_name in folder_names:
            status_label.config(text="Processing...")
            folder_path = os.path.join(search_directory, folder_name)
            print(f"Checking: {folder_path}")  # Debugging line
            if os.path.exists(folder_path):
                status_label.config(text="Processing..")
                target_folder_path = os.path.join(target_directory, folder_name)
                if not os.path.exists(target_folder_path):
                    status_label.config(text="Processing...")
                    os.makedirs(target_folder_path)
                for root, dirs, files in os.walk(folder_path):
                    for file in files:
                        status_label.config(text="Processing..")
                        if any(keyword in file for keyword in keywords):
                            src_file_path = os.path.join(root, file)
                            dest_file_path = os.path.join(target_folder_path, file)
                            shutil.copy2(src_file_path, dest_file_path)
                            csv_writer.writerow([folder_name, src_file_path, dest_file_path])
        status_label.config(text="Processing Completed")


def read_input_file(file_path, id_column=None, last_name_column=None, start_row=1):
    if file_path.lower().endswith('.csv'):
        return read_csv(file_path)
    elif file_path.lower().endswith('.xlsx'):
        return read_excel(file_path, id_column, last_name_column, start_row)



def browse_input_file():
    global input_file
    input_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
    input_file_label.config(text=os.path.basename(input_file))


def browse_search_directory():
    global directory_to_search
    directory_to_search = filedialog.askdirectory()
    search_directory_label.config(text=directory_to_search)


def browse_destination_folder():
    global destination_folder
    global output_csv

    destination_folder = filedialog.askdirectory()
    destination_folder_label.config(text=destination_folder)

    output_csv = os.path.join(destination_folder, "notFound.csv")

def start_processing():
    print(f"Input file: {input_file}")
    print(f"Search directory: {directory_to_search}")
    print(f"Destination directory: {destination_folder}")
    id_column = column_to_index(id_entry.get())
    last_name_column = column_to_index(last_name_entry.get())
    start_row = int(start_row_entry.get())
    folder_names = read_input_file(input_file, id_column, last_name_column, start_row)
    print(f"Folder names: {folder_names}")
    not_found_folders = find_nonexistent_folders(folder_names, directory_to_search)
    print(f"Not found folders: {not_found_folders}")
    write_not_found_csv(not_found_folders, output_csv)
    copy_files_with_keywords(folder_names, directory_to_search, destination_folder, keywords)




# GUI setup
root = Tk()
root.title("Consumer File Verifier")
root.eval('tk::PlaceWindow . center')
root.geometry("400x420")  # Set the default window size to 400x300
root.configure(background='#3c444a')


start_row_label = Label(root, text="Start row:")
start_row_label.pack(pady=5)
start_row_entry = Entry(root, width=5)
start_row_entry.insert(0, "13")  # setting the default value to 13
start_row_entry.pack(pady=5)

id_label = Label(root, text="Consumer_ID column:")
id_label.pack(pady=5)
id_entry = Entry(root, width=5)
id_entry.insert(0, "A")
id_entry.pack(pady=5)

last_name_label = Label(root, text="Last Name column:")
last_name_label.pack(pady=5)
last_name_entry = Entry(root, width=5)
last_name_entry.insert(0, "C")
last_name_entry.pack(pady=5)

input_file_button = Button(root, text="       Select State Report       ", command=browse_input_file, bg='#ed7d2d', fg='#302c29')
input_file_button.pack(pady=5)
input_file_button.pack()

input_file_label = Label(root, text="", background='#3c444a', fg='#ffffff')
input_file_label.pack()

search_directory_button = Button(root, text=" Select Consumers directory ", command=browse_search_directory)
search_directory_button.pack(pady=5)
search_directory_button.pack()

search_directory_label = Label(root, text=directory_to_search, background='#3c444a', fg='#ffffff')
search_directory_label.pack()

destination_folder_button = Button(root, text="    Select destination folder    ", command=browse_destination_folder)
destination_folder_button.pack(pady=5)
destination_folder_button.pack()

destination_folder_label = Label(root, text=destination_folder, background='#3c444a', fg='#ffffff')
destination_folder_label.pack()

start_button = Button(root, text=" Begin Mullinizing Files ", command=start_processing, bg='#08ffb1', fg='#515152',
                      font=("Verdana", 12))
start_button.pack(pady=5)

status_label = Label(root, text="", background='#3c444a', fg='#ffffff')
status_label.pack()

root.mainloop()
